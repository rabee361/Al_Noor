from django.shortcuts import render , HttpResponse , redirect
from base.models import *
from base.resources import PilgrimResource , RegistrationResource
from .forms import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate , login , logout
import pandas as pd
from django.db import transaction
from base.utils.notifications import send_event_notification , send_task_notification
from django.db.models import Q
from django.core.paginator import Paginator
from django.views import View
from django.core.cache import cache
from datetime import datetime, timedelta
from django.utils import timezone
from base.models import FormSubmission
from django.views.generic.edit import CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from openpyxl import load_workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl 
import time
from django.contrib.auth.hashers import make_password

login_decorator = login_required(login_url='login')


def login_user(request):
    if request.method == 'POST':
        phonenumber = request.POST['phonenumber']
        password = request.POST['password']
        user = authenticate(request,username=phonenumber, password=password)
        if user is not None:
            login(request,user)
            return redirect('main_dashboard')

    return render(request , 'admin_panel/login.html')


@login_decorator
def logout_user(request):
    logout(request)
    return redirect('login')


def change_password(request,user_id):
    if request.method == 'POST':
            password1 = request.POST['password1']
            password2 = request.POST['password2']
            if password1 and password2:
                user = CustomUser.objects.get(id=user_id)
                user.set_password(password1)
                user.save()
                return redirect('main_dashboard') 
        
    return render(request, 'admin_panel/change_password.html')

@login_decorator
def main_dashboard(request):
    total_pilgrims = Pilgrim.objects.select_related('user').filter(user__is_deleted=False).count()
    total_employees = Employee.objects.select_related('user').count()
    total_managers = Management.objects.select_related('user').count()
    total_guides = Guide.objects.select_related('user').count()
    total_forms = Registration.objects.filter(is_deleted=False).count()
    new_forms = Registration.objects.filter(is_deleted=False , createdAt__gte=timezone.now() - timedelta(days=1)).count()
    total_tasks = Task.objects.count()
    completed_tasks = Task.objects.filter(completed=True).count()
    total_notes = Note.objects.count()
    
    context = {
    'total_pilgrims' : total_pilgrims,
    'total_employees' : total_employees,
    'total_managers' : total_managers,
    'total_guides' : total_guides,
    'total_forms' : total_forms,
    'new_forms' : new_forms,
    'total_tasks' : total_tasks,
    'completed_tasks' : completed_tasks,
    'total_notes' : total_notes,
    }
    return render(request , 'admin_panel/dashboard.html' , context)

@login_decorator
def steps(request):
    context = {

    }
    return render(request , 'admin_panel/steps/steps.html' , context=context)

@login_decorator
def registration_forms(request):
    q = request.GET.get('q') or ''
    forms = Registration.objects.filter(Q(first_name__startswith = q) & Q(is_deleted=False)).order_by('-id')

    paginator = Paginator(forms, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'forms': page_obj,
    }
    # If it's an HTMX request, return only the partial template
    if request.htmx:
        return render(request, 'admin_panel/partials/registration_forms_partial.html', context)
        
    # Otherwise return the full template
    return render(request, 'admin_panel/registration/registration_forms.html', context)


@login_decorator
def add_register_form(request):
    form = NewRegisterForm()


    if request.method == 'POST':
        form = NewRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('registration_forms')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/registration/add_form.html', context)


@login_decorator
def update_register_form(request,form_id):
    register_form = Registration.objects.get(id=form_id)
    form = UpdateRegisterForm(instance=register_form)

    if request.method == 'POST':
        form = UpdateRegisterForm(request.POST,instance=register_form)
        if form.is_valid():
            form.save()
            return redirect('registration_forms')
        print(form.errors)

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/registration/add_form.html', context)


@login_decorator
def delete_register_form(request,form_id):
    form = Registration.objects.get(id=form_id)
    form.is_deleted=True
    form.save()
    return redirect('registration_forms')


@login_decorator
def delete_all_forms(request):
    Registration.objects.all().update(is_deleted=True)
    return redirect('registration_forms')


@login_decorator
def pilgrims_list(request):
    q = request.GET.get('q') or ''
    pilgrims = Pilgrim.objects.select_related('user').filter(Q(first_name__startswith = q) & Q( user__is_deleted=False)).order_by('-id')

    paginator = Paginator(pilgrims, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'pilgrims': page_obj,
    }

    if request.htmx:
        return render(request, 'admin_panel/partials/pilgrims_partial.html', context)
    return render(request, 'admin_panel/users/pilgrims/pilgrims_list.html', context)


@login_decorator
def update_pilgrim(request,pilgrim_id):
        pilgrim = Pilgrim.objects.get(id=pilgrim_id)
        user = CustomUser.objects.get(id=pilgrim.user.id)
        form = UpdatePilgrimForm(instance=pilgrim)

        if request.method == 'POST':
            pilgrim_form = UpdatePilgrimForm(request.POST,request.FILES,instance=pilgrim)
            print(pilgrim_form.errors)
            if pilgrim_form.is_valid():
                pilgrim = pilgrim_form.save(commit=False)
                user.get_notifications = pilgrim_form.cleaned_data['get_notifications']
                user.phonenumber = pilgrim_form.cleaned_data['phonenumber']
                user.username = request.POST['first_name'] + ' ' + request.POST['father_name'] + ' ' + request.POST['grand_father'] + ' ' + request.POST['last_name']
                image = request.FILES.get('pilgrim_image')
                if image:
                    user.image = image
                user.save()
                logo = request.FILES.get('image')
                if logo:
                    pilgrim.company_logo = logo


                pilgrim.save()

            return redirect('pilgrims')
        
        data = []
        steps = HaJStepsPilgrim.objects.filter(pilgrim=pilgrim).values('haj_step__name')

        total_steps = HajSteps.objects.all()

        for step in total_steps:
            if HaJStepsPilgrim.objects.filter(Q(pilgrim=pilgrim) & Q(haj_step=step.id)).exists():
                data.append({
                    'name':step.name,
                    'rank':step.rank,
                    'completed':True,
                })
            else:
                data.append({
                    'name':step.name,
                    'rank':step.rank,
                    'completed':False,
                })

        try:
            pilgrim_image = pilgrim.user.image.url
        except:
            pilgrim_image = ' '

        context = { 'form': form,
                    'pilgrim_image':pilgrim_image,
                    'pilgrim_id': user.id, 
                    'company_logo':pilgrim.company_logo.url,
                    'pilgrim_steps':data
                    }

        return render(request, 'admin_panel/users/pilgrims/update_pilgrim.html', context=context)


@login_decorator
@transaction.atomic
def add_pilgrim(request):
    form = NewPilgrim()

    if request.method == 'POST':
        form = NewPilgrim(request.POST, request.FILES)
        if form.is_valid():
            print(form.error_class)
            user = CustomUser.objects.create(
                username=form.cleaned_data['first_name'],
                phonenumber=form.cleaned_data['phonenumber'],
                # email=form.cleaned_data['email'],
                get_notifications=form.cleaned_data['get_notifications'],
                user_type = 'حاج'
            )
            user.set_password(form.cleaned_data['password'])
            if request.POST['image']:
                user.image = request.POST['image']
                user.save()

            Chat.objects.create(user=user , chat_type='guide')
            Chat.objects.create(user=user , chat_type='manager')

            pilgrim = Pilgrim.objects.create(
                user=user,
                first_name=form.cleaned_data['first_name'],
                father_name=form.cleaned_data['father_name'],
                last_name=form.cleaned_data['last_name'],
                grand_father=form.cleaned_data['grand_father'],
                registeration_id=form.cleaned_data['registeration_id'],
                phonenumber=form.cleaned_data['phonenumber'],
                hotel=form.cleaned_data['hotel'],
                hotel_address=form.cleaned_data['hotel_address'],
                room_num=form.cleaned_data['room_num'],
                gate_num=form.cleaned_data['gate_num'],
                flight_num=form.cleaned_data['flight_num'],
                flight_date=form.cleaned_data['flight_date'],
                flight_company = form.cleaned_data['flight_company'],
                company_logo = form.cleaned_data['company_logo'],
                from_city=form.cleaned_data['from_city'],
                to_city=form.cleaned_data['to_city'],
                birthday=request.POST['birthday'],
                duration=form.cleaned_data['duration'],
                boarding_time=request.POST['boarding_time'],
                arrival=request.POST['arrival'],
                departure=request.POST['departure'],
            )
            if request.POST['guide']:
                guide = Guide.objects.get(id=request.POST['guide'])
                pilgrim.guide = guide
                pilgrim.save()

            return redirect('pilgrims')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/users/pilgrims/add_pilgrim.html', context)



@login_decorator
def delete_pilgrim(request,pilgrim_id):
    pilgrim = Pilgrim.objects.get(id=pilgrim_id)
    user = CustomUser.objects.get(id=pilgrim.user.id)
    user.is_deleted = True
    user.save()
    return redirect('pilgrims')


@login_decorator
def delete_all_pilgrims(request):
    users = CustomUser.objects.filter(user_type='حاج')
    users.update(is_deleted=True)
    return redirect('pilgrims')


@login_decorator
def managers_list(request):
    q = request.GET.get('q') or ''
    managers = Management.objects.filter(user__first_name__startswith = q).order_by('-id')

    paginator = Paginator(managers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'managers':page_obj,
    }
    return render(request , 'admin_panel/users/managers/managers_list.html' , context)


@login_decorator
def add_manager(request):
    form = NewManager()

    if request.method == 'POST':
        form = NewManager(request.POST, request.FILES)
        if form.is_valid():
            print(form.cleaned_data['password1'])
            user = CustomUser.objects.create(
                username=form.cleaned_data['username'],
                phonenumber=form.cleaned_data['phonenumber'],
                email=form.cleaned_data['email'],
                get_notifications=form.cleaned_data['get_notifications'],
                user_type = 'اداري'
            )
            user.set_password(form.cleaned_data['password1'])
            user.save()
            if form.cleaned_data['image']:
                user.image = form.cleaned_data['image']
            
            Management.objects.create(user=user)
            return redirect('managers')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/users/managers/manager_form.html' , context)


@login_decorator
def update_manager(request,manager_id):
    manager = Management.objects.get(id=manager_id)
    user = manager.user
    form = UpdateManager(instance=user)

    if request.method == 'POST':
        form = UpdateManager(request.POST,request.FILES,instance=manager)
        if form.is_valid():
            image=request.FILES.get('image')
            if image:
                user.image = request.FILES.get('image')
            else:
                user.image = manager.user.image

            user.phonenumber = form.cleaned_data['phonenumber']
            user.get_notifications = form.cleaned_data['get_notifications']
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.save()
            return redirect('managers')

    user_image = request.user.image.url
    manager_image = manager.user.image.url
    username = request.user.username
    
    context = {
        'form': form,
        'manager_image': manager_image,
        'manager_id': user.id,
    }
    return render(request, 'admin_panel/users/managers/update_manager.html', context)


@login_decorator
def delete_manager(request,manager_id):
    manager = Management.objects.get(id=manager_id)
    user = CustomUser.objects.get(id=manager.user.id)
    user.delete()
    return redirect("managers")


@login_decorator
def task_list(request):
    q = request.GET.get('q') or ''
    tasks = Task.objects.filter(employee__user__username__startswith = q).order_by('-id')

    paginator = Paginator(tasks, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'tasks':page_obj,
    }

    if request.htmx:
        return render(request , 'admin_panel/partials/tasks_partial.html' , context)
    return render(request , 'admin_panel/tasks/tasks.html' , context)


@login_decorator
def add_task(request):
    form = NewTask()

    if request.method == 'POST':
        form = NewTask(request.POST)
        if form.is_valid():
            form.save(commit=False)
            send_task_notification(employee=form.cleaned_data['employee'],
                                   title=form.cleaned_data['title'],
                                   content=form.cleaned_data['content'])
            form.save()
            return redirect('tasks')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/tasks/task_form.html', context)


@login_decorator
def update_task(request,task_id):
    task = Task.objects.get(id=task_id)
    form = NewTask(instance=task)

    if request.method == 'POST':
        form = NewTask(request.POST, instance=task)
        if form.is_valid():
            form.save()
            return redirect('tasks')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/tasks/task_form.html', context)


@login_decorator
def delete_task(request,task_id):
    Task.objects.get(id=task_id).delete()

    return redirect('tasks')


@login_decorator
def notes_list(request):
    q = request.GET.get('q') or ''
    notes = Note.objects.filter(pilgrim__user__username__startswith = q).order_by('-id')

    paginator = Paginator(notes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'notes':page_obj,
    }
    if request.htmx:
        return render(request , 'admin_panel/partials/notes_partial.html' , context)
    return render(request , 'admin_panel/notes/notes.html' , context)


@login_decorator
def add_note(request):
    form = NewNote()

    if request.method == 'POST':
        form = NewNote(request.POST)
        if form.is_valid():
            form.save()
            return redirect('notes')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/notes/add_note.html', context)


@login_decorator
def update_note(request,note_id):
    note = Note.objects.get(id=note_id)
    form = NewNote(instance=note)

    if request.method == 'POST':
        form = NewNote(request.POST, instance=note)
        if form.is_valid():
            form.save()
            return redirect('notes')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/notes/add_note.html', context)


@login_decorator
def delete_note(request,note_id):
    Note.objects.get(id=note_id).delete()

    return redirect('notes')


@login_decorator
def employees_list(request):
    q = request.GET.get('q') or ''

    employees = Employee.objects.filter(user__username__startswith=q)

    paginator = Paginator(employees, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'employees': page_obj,
    }
    return render(request , 'admin_panel/users/employees/employees_list.html' , context)


@login_decorator
def add_employee(request):
    form = NewEmployee()

    if request.method == 'POST':
        form = NewEmployee(request.POST, request.FILES)
        if form.is_valid():
            user = CustomUser.objects.create(
                username=form.cleaned_data['username'],
                phonenumber=form.cleaned_data['phonenumber'],
                email=form.cleaned_data['email'],
                get_notifications=form.cleaned_data['get_notifications'],
                user_type = 'موظف'
            )
            user.set_password(form.cleaned_data['password1'])
            user.save()
            if form.cleaned_data['image']:
                user.image = form.cleaned_data['image']

            Employee.objects.create(user=user)
            return redirect('employees')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/users/employees/add_employee.html' , context)


@login_decorator
def update_employee(request,employee_id):
    employee = Employee.objects.get(id=employee_id)
    user = employee.user
    form = UpdateEmployee(instance=user)

    if request.method == 'POST':
        form = UpdateEmployee(request.POST,request.FILES,instance=employee)
        if form.is_valid():
            image=request.FILES.get('image')
            if image:
                user.image = request.FILES.get('image')
            else:
                user.image = employee.user.image


            user.get_notifications = form.cleaned_data['get_notifications']
            user.phonenumber = form.cleaned_data['phonenumber']
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.save()
            return redirect('employees')
        
    employee_image = request.user.image.url
    employee_image = employee.user.image.url
    username = request.user.username
    
    context = {
        'form': form,
        'employee_id': user.id,
        'employee_image': employee_image,
    }
    return render(request, 'admin_panel/users/employees/update_employee.html', context)


@login_decorator
def delete_employee(request,employee_id):
    employee = Employee.objects.get(id=employee_id)
    user = CustomUser.objects.get(id=employee.user.id)
    user.delete()
    return redirect('employees')


@login_decorator
def guides_list(request):
    q = request.GET.get('q') or ''

    guides = Guide.objects.filter(user__username__startswith=q).order_by('-id')

    paginator = Paginator(guides, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'guides':page_obj,
    }
    return render(request , 'admin_panel/users/guides/guides_list.html' , context)


@login_decorator
def add_guide(request):
    form = NewGuide()

    if request.method == 'POST':
        form = NewGuide(request.POST, request.FILES)
        if form.is_valid():
            print(form.cleaned_data['image'])
            user = CustomUser.objects.create(
                username=form.cleaned_data['username'],
                phonenumber=form.cleaned_data['phonenumber'],
                email=form.cleaned_data['email'],
                get_notifications=form.cleaned_data['get_notifications'],
                user_type = 'مرشد'
            )
            user.set_password(form.cleaned_data['password1'])
            user.save()
            if form.cleaned_data['image']:
                user.image = form.cleaned_data['image'] 

            Guide.objects.create(user=user)
            return redirect('guides')
    context = {
        'form' : form,

    }
    return render(request , 'admin_panel/users/guides/add_guide.html' , context)


@login_decorator
def update_guide(request,guide_id):
    guide = Guide.objects.get(id=guide_id)
    user = guide.user
    form = UpdateGuide(instance=user)

    if request.method == 'POST':
        form = UpdateGuide(request.POST,request.FILES,instance=guide)
        if form.is_valid():
            image=request.FILES.get('image')
            if image:
                user.image = request.FILES.get('image')
            else:
                user.image = guide.user.image

            user.phonenumber = form.cleaned_data['phonenumber']
            user.get_notifications = form.cleaned_data['get_notifications']
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.save()
            return redirect('guides')

    guide_image = guide.user.image.url
    
    context = {
        'form': form,
        'guide_image': guide_image,
        'guide_id': user.id,
    }
    return render(request, 'admin_panel/users/guides/update_guide.html', context)


@login_decorator
def delete_guide(request,guide_id):
    guide = Guide.objects.get(id=guide_id)
    user = CustomUser.objects.get(id=guide.user.id)
    user.delete()
    return redirect('guides')


@login_decorator
def export_pilgram(request):
    pilgrim_resource = PilgrimResource()
    dataset = pilgrim_resource.export()
    response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="pilgrims.xlsx"'
    return response


@login_decorator
def export_forms(request):
    pilgrim_resource = RegistrationResource()
    dataset = pilgrim_resource.export()
    response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="register_forms.xlsx"'
    return response



@login_decorator
@transaction.atomic
def import_pilgrim(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES['file']
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # Assuming the first row is the header
            headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}
            
            # Initialize counters and timer
            start_time = time.time()
            total_rows = 0
            processed_rows = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                total_rows += 1
                try:
                    first_name = str(row[headers['الاسم الأول']])
                    last_name = str(row[headers['العائلة']])
                    father_name = str(row[headers['اسم الأب']])
                    grand_father = str(row[headers['اسم الجد']])
                    birth_date = str(row[headers['تاريخ الميلاد - الميلادي فقط']])
                    flight_num = str(row[headers['رقم الرحلة']])
                    card_num = str(row[headers['رقم الهوية']])
                    phone_number = str(row[headers['رقم الجوال']])
                    flight_date = str(row[headers['تاريخ الرحلة']])
                    from_city = str(row[headers['من المدينة']])
                    to_city = str(row[headers['إلى المدينة']])
                    hotel = str(row[headers['الفندق']])
                    hotel_address = str(row[headers['عنوان الفندق']])
                    room_num = str(row[headers['رقم الغرفة']])
                    gate_num = str(row[headers['رقم البوابة']])
                    flight_company = str(row[headers['شركة الطيران']])
                    arrival_time = str(row[headers['موعد الوصول']])
                    departure_time = str(row[headers['موعد الرحيل']])
                    boarding_time = str(row[headers['وقت الصعود']])

                    # Start timing for time conversion block
                    time_conversion_start = time.time()
                    
                    # Convert time strings to proper datetime format
                    arrival_str = arrival_time.strip()
                    departure_str = departure_time.strip()
                    boarding_str = boarding_time.strip()

                    # Parse times and convert to 24-hour format
                    try:
                        arrival_dt = datetime.strptime(arrival_str, '%I:%M:%S %p')
                        arrival_time = arrival_dt.strftime('%H:%M:%S')
                    except ValueError:
                        arrival_time = '00:00:00'  # Default if parsing fails

                    try:
                        departure_dt = datetime.strptime(departure_str, '%I:%M:%S %p') 
                        departure_time = departure_dt.strftime('%H:%M:%S')
                    except ValueError:
                        departure_time = '00:00:00'

                    try:
                        boarding_dt = datetime.strptime(boarding_str, '%I:%M:%S %p')
                        boarding_time = boarding_dt.strftime('%H:%M:%S') 
                    except ValueError:
                        boarding_time = '00:00:00'

                    # Calculate duration
                    try:
                        diff = departure_dt - arrival_dt
                        formatted_diff = str(timedelta(hours=diff.seconds//3600, minutes=diff.seconds//60%60))
                    except:
                        formatted_diff = '00:00'
                    
                    time_conversion_end = time.time()
                    print(f"Time conversion block took: {time_conversion_end - time_conversion_start:.4f} seconds")

                    # Start timing for user creation/update block
                    user_block_start = time.time()
                    
                    pilgrim_phonenumber = phone_number
                    if not pilgrim_phonenumber:
                        continue

                    pilgrim_username = f"{first_name} {father_name} {grand_father} {last_name}"
                    user, created = CustomUser.objects.update_or_create(
                        phonenumber=pilgrim_phonenumber,
                        defaults={
                            'username': pilgrim_username,
                            'user_type': 'حاج',
                            'first_name': first_name,
                            'last_name': last_name,
                            'is_deleted': False,
                            'password': make_password(card_num)
                        }
                    )
                    if created:                
                        Chat.objects.bulk_create([
                            Chat(user=user, chat_type='guide'),
                            Chat(user=user, chat_type='manager')
                        ])

                    user_block_end = time.time()
                    print(f"User creation/update block took: {user_block_end - user_block_start:.4f} seconds")

                    # Start timing for pilgrim creation/update block
                    pilgrim_block_start = time.time()
                    
                    pilgrim, created = Pilgrim.objects.update_or_create(
                        user=user,
                        phonenumber=pilgrim_phonenumber,
                        defaults={
                            'registeration_id': card_num,
                            'first_name': first_name,
                            'father_name': father_name,
                            'grand_father': grand_father,
                            'last_name': last_name,
                            'birthday': birth_date,
                            'flight_num': flight_num,
                            'flight_date': flight_date,
                            'arrival': arrival_time,
                            'departure': departure_time,
                            'from_city': from_city,
                            'to_city': to_city,
                            'duration': str(formatted_diff),
                            'boarding_time': boarding_time,
                            'gate_num': gate_num,
                            'flight_company': flight_company,
                            'status': True,
                            'hotel': hotel,
                            'hotel_address': hotel_address,
                            'room_num': room_num,
                        }
                    )

                    # Start timing for guide assignment block
                    guide_block_start = time.time()
                    
                    try:
                        guide = Guide.objects.get(id=row[headers['المرشد']])
                        pilgrim.guide = guide
                    except Guide.DoesNotExist:
                        pass

                    if created:
                        pilgrim.save()
                    
                    guide_block_end = time.time()
                    print(f"Guide assignment block took: {guide_block_end - guide_block_start:.4f} seconds")
                    
                    pilgrim_block_end = time.time()
                    print(f"Pilgrim creation/update block took: {pilgrim_block_end - pilgrim_block_start:.4f} seconds")
                    
                    processed_rows += 1

                except Exception as e:
                    print(f"Error processing row {total_rows + 1}: {str(e)}")
                    continue

            # Calculate execution time and statistics
            end_time = time.time()
            execution_time = end_time - start_time
            
            print("="*50)
            print("Import Statistics:")
            print(f"Total rows in file: {total_rows}")
            print(f"Successfully processed rows: {processed_rows}")
            print(f"Failed rows: {total_rows - processed_rows}")
            print(f"Total execution time: {execution_time:.2f} seconds")
            print(f"Average time per record: {(execution_time/total_rows if total_rows > 0 else 0):.2f} seconds")
            print("="*50)

            return redirect('pilgrims')
            
        except Exception as e:
            print(f"Import error: {str(e)}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return HttpResponse(str(e), status=500)
            return redirect('pilgrims')

    return render(request, 'admin_panel/users/pilgrims/import_pilgrims.html')


@login_decorator
def notifications_list(request):
    q = request.GET.get('q') or ''
    notifications = BaseNotification.objects.filter(title__startswith=q).order_by('-id')

    paginator = Paginator(notifications, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'notifications':page_obj,
    }
    return render(request , 'admin_panel/notifications/notifications_list.html' , context)







@login_decorator
def add_notification(request):
    form = NotificationForm()
    user = request.user
    if request.method == 'POST':
        form = NotificationForm(request.POST, request.FILES)
        if form.is_valid():
            title = form.cleaned_data['title']
            content = form.cleaned_data['content']
            send_event_notification(title=title,content=content,sentBy=user)
            return redirect('notifications')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/notifications/add_notification.html' , context)

@login_decorator
def delete_notifications(request,notification_id):
    BaseNotification.objects.get(id=notification_id).delete()
    return redirect('notifications')


@login_decorator
def guidance_posts(request):
    q = request.GET.get('q') or ''
    posts = GuidancePost.objects.filter(title__startswith=q).order_by('rank')
    context = {
        'posts':posts,
    }
    return render(request , 'admin_panel/religious_work/guidance_posts.html' , context)




@login_decorator
def add_guidance_post(request):
    form = GuidancePostForm()
    if request.method == 'POST':
        form = GuidancePostForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('guidance_posts')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/religious_work/add_guidance_post.html' , context)



@login_decorator
def update_guidance_post(request,post_id):
    post = GuidancePost.objects.get(id=post_id)
    form = GuidancePostForm(instance=post)

    if request.method == 'POST':
        form = GuidancePostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            if request.FILES.get('cover'):
                post.cover = request.FILES.get('cover')
                post.save()
            form.save()
            return redirect('guidance_posts')

    context = {
        'form': form,
        'post_image':post.cover.url
    }
    return render(request, 'admin_panel/religious_work/update_guidance_post.html', context)



@login_decorator
def delete_guidance_post(request,post_id):
    GuidancePost.objects.get(id=post_id).delete()
    return redirect('guidance_posts')



@login_decorator
def guidance_categories(request): 
    q = request.GET.get('q') or ''
    categories = GuidanceCategory.objects.filter(name__startswith=q).order_by('-id')
    context = {
        'categories':categories,
    }
    return render(request , 'admin_panel/religious_work/guidance_categories.html' , context)




@login_decorator
def add_guidance_category(request):
    form = GuidanceCategoryForm()
    if request.method == 'POST':
        form = GuidanceCategoryForm(request.POST)
        print(form.errors)
        if form.is_valid():
            form.save()
            return redirect('guidance_categories')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/religious_work/add_guidance_category.html' , context)








@login_decorator
def update_guidance_category(request,category_id):
    category = GuidanceCategory.objects.get(id=category_id)
    form = GuidanceCategoryForm(instance=category)

    if request.method == 'POST':
        form = GuidanceCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('guidance_categories')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/religious_work/add_guidance_category.html', context)





@login_decorator
def delete_guidance_category(request,category_id):
    GuidanceCategory.objects.get(id=category_id).delete()
    return redirect('guidance_categories')








def religious_posts(request):
    q = request.GET.get('q') or ''
    posts = ReligiousPost.objects.filter(title__startswith=q).order_by('rank')
    context = {
        'posts':posts,
    }
    return render(request , 'admin_panel/religious_work/religious_posts.html' , context)






@login_decorator
def add_religious_post(request):
    form = ReligiousPostForm()
    if request.method == 'POST':
        form = ReligiousPostForm(request.POST, request.FILES)
        print(form.errors)
        if form.is_valid():
            form.save()
            return redirect('religious_posts')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/religious_work/add_religious_post.html' , context)









@login_decorator
def update_religious_post(request,post_id):
    post = ReligiousPost.objects.get(id=post_id)
    form = ReligiousPostForm(instance=post)

    if request.method == 'POST':
        form = ReligiousPostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            if request.FILES.get('cover'):
                post.cover = request.FILES.get('cover')
                post.save()
            form.save()
            return redirect('religious_posts')

    context = {
        'form': form,
        'post_image':post.cover.url
    }
    return render(request, 'admin_panel/religious_work/update_religious_post.html', context)





@login_decorator
def delete_religious_post(request,post_id):
    ReligiousPost.objects.get(id=post_id).delete()
    return redirect('religious_posts')





@login_decorator
def religious_categories(request): 
    q = request.GET.get('q') or ''
    categories = ReligiousCategory.objects.filter(name__startswith=q).order_by('-id')
    context = {
        'categories':categories,
    }
    return render(request , 'admin_panel/religious_work/religious_categories.html' , context)






@login_decorator
def add_religious_category(request):
    form = ReligiousCategoryForm()

    if request.method == 'POST':
        form = ReligiousCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('religious_categories')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/religious_work/add_religious_category.html' , context)





@login_decorator
def update_religious_category(request,category_id):
    category = ReligiousCategory.objects.get(id=category_id)
    form = ReligiousCategoryForm(instance=category)

    if request.method == 'POST':
        form = ReligiousCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('religious_categories')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/religious_work/add_religious_category.html', context)





@login_decorator
def delete_religious_category(request,category_id):
    ReligiousCategory.objects.get(id=category_id).delete()
    return redirect('religious_categories')





@login_decorator
def update_religious_post(request,post_id):
    post = ReligiousPost.objects.get(id=post_id)
    form = ReligiousPostForm(instance=post)

    if request.method == 'POST':
        form = ReligiousPostForm(request.POST, instance=post)
        if form.is_valid():
            form.save()
            return redirect('religious_posts')

    context = {
        'form': form,
        'post_image':post.cover.url
    }
    return render(request, 'admin_panel/religious_work/update_religious_post.html', context)





@login_decorator
def delete_religious_post(request,post_id):
    ReligiousPost.objects.get(id=post_id).delete()
    return redirect('religious_posts')


def stream_types(request):
    q = request.GET.get('q') or ''
    types = LiveStreamCategory.objects.filter(name__startswith=q).order_by('-id')
    context = {
        'types':types,
    }
    return render(request , 'admin_panel/live_stream/stream_types.html' , context)

def add_stream_type(request):
    form = LiveStreamTypeForm()
    if request.method == 'POST':
        form = LiveStreamTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stream_types')
    context = {
        'form': form,
    }
    return render(request, 'admin_panel/live_stream/add_stream_type.html', context)

def update_stream_type(request,id):
    type = LiveStreamCategory.objects.get(id=id)
    form = LiveStreamTypeForm(instance=type)
    if request.method == 'POST':
        form = LiveStreamTypeForm(request.POST, instance=type)
        if form.is_valid():
            form.save()
            return redirect('stream_types')
    context = {
        'form': form,
    }
    return render(request, 'admin_panel/live_stream/add_stream_type.html', context)
        
    
def delete_stream_type(request,id):
    LiveStreamCategory.objects.get(id=id).delete()
    return redirect('stream_types')


def live_streams(request):
    streams = LiveStream.objects.all()
    context = {
        'streams':streams,
    }
    return render(request , 'admin_panel/live_stream/live_streams.html' , context)

def add_live_stream(request):
    form = LiveStreamForm()
    print(request.POST)
    if request.method == 'POST':
        form = LiveStreamForm(request.POST , request.FILES)
        if form.is_valid():
            form.save()
            return redirect('live_streams')
        print(form.errors)
    context = {
        'form': form,
    }
    return render(request, 'admin_panel/live_stream/add_live_stream.html', context)

def update_live_stream(request,id):
    stream = LiveStream.objects.get(id=id)
    form = LiveStreamForm(instance=stream)

    if request.method == 'POST':
        form = LiveStreamForm(request.POST, request.FILES, instance=stream)
        if form.is_valid():
            if request.FILES.get('cover'):
                stream.cover = request.FILES.get('cover')
                stream.save()
            form.save()
            return redirect('live_streams')

    context = {
        'form': form,
        'stream_cover':stream.cover.url
    }
    return render(request, 'admin_panel/live_stream/update_live_stream.html', context)

def delete_live_stream(request,id):
    LiveStream.objects.get(id=id).delete()
    return redirect('live_streams')


@login_decorator
def steps_list(request):
    q = request.GET.get('q') or ''
    steps = HajSteps.objects.filter(name__startswith=q).order_by('-id')
    context = {
        'steps':steps,
    }
    return render(request , 'admin_panel/steps/steps.html' , context)



@login_decorator
def pilgrim_steps(request):
    q = request.GET.get('q', '')
    page = request.GET.get('page', 1)
    
    # Get all pilgrims first
    pilgrims = Pilgrim.objects.select_related('user').filter(
        user__username__startswith=q,
        user__is_deleted=False
    ).order_by('user__username')
    
    # Get all steps in one query
    steps = HajSteps.objects.all().order_by('rank')
    
    # Get completed steps for all filtered pilgrims in one query
    pilgrim_ids = list(pilgrims.values_list('id', flat=True))
    completed_steps = set(
        HaJStepsPilgrim.objects.filter(
            pilgrim_id__in=pilgrim_ids
        ).values_list('pilgrim_id', 'haj_step_id')
    )
    
    # Build the data structure efficiently
    data = []
    for pilgrim in pilgrims:
        for step in steps:
            data.append({
                'name': step.name,
                'pilgrim': pilgrim.user.username,
                'completed': (pilgrim.id, step.id) in completed_steps
            })
    
    # Paginate the final data
    paginator = Paginator(data, 10)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'steps': page_obj,
        'page_obj': page_obj,
        'total_steps': len(steps),
        'search_query': q
    }
    
    if request.htmx:
        return render(request, 'admin_panel/partials/pilgrim_steps_partial.html', context)
    return render(request, 'admin_panel/steps/pilgrim_steps.html', context)



@login_decorator
def add_step(request):
    form = StepForm()

    if request.method == 'POST':
        form = StepForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('steps')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/steps/add_step.html' , context)





@login_decorator
def update_step(request,step_id):
    step = HajSteps.objects.get(id=step_id)
    form = StepForm(instance=step)

    if request.method == 'POST':
        form = StepForm(request.POST, instance=step)
        if form.is_valid():
            form.save()
            return redirect('steps')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/steps/add_step.html', context)





@login_decorator
def delete_step(request,step_id):
    HajSteps.objects.get(id=step_id).delete()
    return redirect('steps')









@login_decorator
def secondary_steps_list(request):
    q = request.GET.get('q') or ''
    steps = SecondarySteps.objects.filter(name__startswith=q).order_by('-id')
    context = {
        'steps':steps,
    }
    return render(request , 'admin_panel/steps/secondary_steps.html' , context)







@login_decorator
def add_secondary_step(request):
    form = SecondaryStepForm()

    if request.method == 'POST':
        form = SecondaryStepForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('secondary_steps')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/steps/add_secondary_step.html' , context)







@login_decorator
def update_secondary_step(request,step_id):
    step = SecondarySteps.objects.get(id=step_id)
    form = SecondaryStepForm(instance=step)

    if request.method == 'POST':
        form = SecondaryStepForm(request.POST, instance=step)
        if form.is_valid():
            form.save()
            return redirect('secondary_steps')

    context = {
        'form': form,
    }
    return render(request, 'admin_panel/steps/add_secondary_step.html', context)








@login_decorator
def delete_secondary_step(request,step_id):
    SecondarySteps.objects.get(id=step_id).delete()
    return redirect('secondary_steps')




@login_decorator
def my_account(request):
    user = request.user
    form = UpdateUser(instance=user)

    if request.method == 'POST':
        form = UpdateUser(request.POST,request.FILES,instance=user)
        if form.is_valid():
            user = CustomUser.objects.get(
                phonenumber=form.cleaned_data['phonenumber'],
            )
            image=request.FILES.get('image')
            if image:
                user.image = request.FILES.get('image')
            else:
                user.image = user.image


            user.get_notifications = form.cleaned_data['get_notifications']
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.save()
            return redirect('employees')

    
    context = {
        'form': form,
        'user_id': user.id,
    }
    
    return render(request , 'my_account.html' , context=context)





@login_decorator
def admins_list(request):
    q = request.GET.get('q') or ''
    admins = CustomUser.objects.filter(is_superuser=True,username__startswith=q).order_by('-id')
    context = {
        'admins':admins,
    }
    return render(request , 'admin_panel/users/admin/admin_list.html' , context)



@login_decorator
def update_admin(request,admin_id):
    admin = CustomUser.objects.get(id=admin_id)
    form = UpdateAdmin(instance=admin)

    if request.method == 'POST':
        form = UpdateAdmin(request.POST,request.FILES,instance=admin)
        if form.is_valid():
            image=request.FILES.get('image')
            if image:
                admin.image = request.FILES.get('image')
            else:
                admin.image = admin.image

            admin.username = form.cleaned_data['username']
            admin.email = form.cleaned_data['email']
            admin.phonenumber = form.cleaned_data['phonenumber']
            admin.save()
            return redirect('admins')
        
    admin_image = request.user.image.url
    admin_image = admin.image.url
    username = request.user.username
    
    context = {
        'form': form,
        'admin_image': admin_image,
        'admin_id': admin.id,
    }
    return render(request, 'admin_panel/users/admin/update_admin.html', context)




@login_decorator
def delete_admin(request,admin_id):
    CustomUser.objects.get(id=admin_id).delete()
    return redirect('admins')



@login_decorator
def add_admin(request):
    form = NewAdmin()

    if request.method == 'POST':
        form = NewAdmin(request.POST, request.FILES)
        if form.is_valid():
            user = CustomUser.objects.create(
                username=form.cleaned_data['username'],
                phonenumber=form.cleaned_data['phonenumber'],
                email=form.cleaned_data['email'],
                get_notifications=form.cleaned_data['get_notifications'],
                user_type = 'اداري',
                is_superuser = True,
                is_staff = True
            )
            user.set_password(form.cleaned_data['password1'])
            user.save()
            if form.cleaned_data['image']:
                user.image = form.cleaned_data['image']
            
            return redirect('main_dashboard')
    context = {
        'form' : form,
    }
    return render(request , 'admin_panel/users/admin/add_admin.html' , context)




def terms(request):
    terms = TermsAndConditions.objects.first()
    context = {
        'terms':terms,
    }
    return render(request , 'admin_panel/about/terms_and_privacy.html' , context)



@login_decorator
def update_terms(request):
    terms = TermsAndConditions.objects.first()
    form = TermsForm(instance=terms)
    if request.method == 'POST':
        form = TermsForm(request.POST,instance=terms)
        if form.is_valid():
            form.save()
            return redirect('terms')
    context = {
        'form':form,
    }
    return render(request , 'admin_panel/about/update_terms.html' , context)




class TermsandConditionsView(View):
    def get(self,request):
        terms = TermsAndConditions.objects.first()
        context = {
            'terms':terms.terms,
            'privacy':terms.privacy,
        }
        return render(request , 'terms_and_conditions.html' , context)




class PilgrimFormView(View):
    MAX_SUBMISSIONS = 10
    
    def get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

    def get_submission_count(self, ip):
        # Get submissions from last 24 hours
        time_threshold = timezone.now() - timedelta(hours=24)
        return FormSubmission.objects.filter(
            ip_address=ip,
            submitted_at__gte=time_threshold
        ).count()

    def add_submission(self, ip):
        FormSubmission.objects.create(ip_address=ip)

    def get(self, request):
        form = NewRegisterForm()
        success = request.session.pop('form_success', False)
        
        # Check submission count
        ip = self.get_client_ip(request)
        submission_count = self.get_submission_count(ip)
        
        if submission_count >= self.MAX_SUBMISSIONS:
            return render(request, 'pilgrim_form/form.html', {
                'form': form,
                'submission_limit_reached': True,
                'max_submissions': self.MAX_SUBMISSIONS
            })
        
        return render(request, 'pilgrim_form/form.html', {
            'form': form, 
            'success': success
        })
    
    def post(self, request):
        # Check if this is a duplicate submission
        form_id = request.session.get('last_form_id')
        current_form_id = request.POST.get('form_id', None)
        
        if form_id and form_id == current_form_id:
            # This is a duplicate submission, redirect to fresh form
            return redirect('form')

        # Check submission limit
        ip = self.get_client_ip(request)
        submission_count = self.get_submission_count(ip)
        
        # Block if limit is reached
        if submission_count >= self.MAX_SUBMISSIONS:
            return render(request, 'pilgrim_form/form.html', {
                'form': NewRegisterForm(),
                'submission_limit_reached': True,
                'max_submissions': self.MAX_SUBMISSIONS
            })
            
        form = NewRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            # Record the submission
            self.add_submission(ip)
            # Store the form ID in session
            request.session['last_form_id'] = current_form_id
            request.session['form_success'] = True
            
            # Check if this was the final submission
            new_submission_count = self.get_submission_count(ip)
            if new_submission_count >= self.MAX_SUBMISSIONS:
                return render(request, 'pilgrim_form/form.html', {
                    'form': NewRegisterForm(),
                    'success': True,
                    'submission_limit_reached': True,
                    'max_submissions': self.MAX_SUBMISSIONS
                })
            
            return redirect('form')
        else:
            # If form is invalid, preserve the data and show errors
            return render(request, 'pilgrim_form/form.html', {
                'form': form,  # This will contain the submitted data and errors
                'form_data': request.POST,  # Pass the POST data to preserve values
                'active_tab': request.POST.get('active_tab', '1')  # Preserve active tab
            })





class LandinPageView(View):
    def get(self, request):
        return render(request , 'pilgrim_form/landing.html')


class AddPilgrimView(LoginRequiredMixin, CreateView):
    template_name = 'admin_panel/users/pilgrims/add_pilgrim.html'
    form_class = PilgrimCreationForm
    success_url = reverse_lazy('pilgrims')
    login_url = 'login'

    @transaction.atomic
    def form_valid(self, form):
        # Create CustomUser
        user = CustomUser.objects.create(
            username=form.cleaned_data['first_name'],
            phonenumber=form.cleaned_data['phonenumber'],
            get_notifications=form.cleaned_data['get_notifications'],
            user_type='حاج'
        )
        user.set_password(form.cleaned_data['password'])
        
        if form.cleaned_data.get('image'):
            user.image = form.cleaned_data['image']
        user.save()

        # Create Chat objects
        Chat.objects.create(user=user, chat_type='guide')
        Chat.objects.create(user=user, chat_type='manager')

        # Create Pilgrim
        pilgrim = form.save(commit=False)
        pilgrim.user = user
        
        if form.cleaned_data.get('guide'):
            pilgrim.guide = form.cleaned_data['guide']
            
        pilgrim.save()
        
        return super().form_valid(form)
    


class handler404(View):
    def get(self, request):
        return render(request, '404.html')

class handler500(View):
    def get(self, request):
        return render(request, '500.html')
